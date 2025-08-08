import streamlit as st
import pandas as pd
import random
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# --- CORE LOGIC FUNCTIONS ---

def find_lowest_competencies(row, competencies):
    """Identifies the two lowest scoring competencies for a candidate."""
    scores = {comp: row[comp] for comp in competencies if pd.notna(row[comp])}
    sorted_competencies = sorted(scores.items(), key=lambda item: item[1])
    
    lowest_1 = sorted_competencies[0][0] if len(sorted_competencies) > 0 else None
    lowest_2 = sorted_competencies[1][0] if len(sorted_competencies) > 1 else None
    
    return lowest_1, lowest_2

def get_random_tips(repo_df, competency):
    """Gets two random tips (70% and 20%) for a given competency."""
    # This line handles blank cells in the Competency Name column.
    repo_df.dropna(subset=['Competency Name'], inplace=True)
    
    # --- FIXED: Corrected the line that caused the error ---
    # The '.str' was removed from 'competency.strip().lower()'
    comp_df = repo_df[repo_df['Competency Name'].str.strip().str.lower() == competency.strip().lower()]
    
    tips_70 = comp_df['70% Development Tips'].dropna().tolist()
    tips_20 = comp_df['20% Development Tips'].dropna().tolist()
    
    tip_70 = random.choice(tips_70) if tips_70 else "N/A"
    tip_20 = random.choice(tips_20) if tips_20 else "N/A"
    
    return tip_70, tip_20

def generate_formatted_excel(results):
    """Creates a stylized Excel workbook with one sheet per candidate."""
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active) 

    title_font = Font(name='Calibri', size=18, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True)
    body_font = Font(name='Calibri', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for res in results:
        ws = wb.create_sheet(title=res['Candidate Name'][:30])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        ws.merge_cells('A1:B1')
        ws['A1'].value = "Development Plan"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 30
        ws['A3'] = "Candidate Name:"
        ws['A3'].font = header_font
        ws['B3'] = res['Candidate Name']
        ws['B3'].font = body_font
        ws['A4'] = "Level:"
        ws['A4'].font = header_font
        ws['B4'] = res['Level']
        ws['B4'].font = body_font
        ws.merge_cells('A6:B6')
        ws['A6'].value = f"Focus Area 1: {res['Lowest Competency 1']}"
        ws['A6'].font = header_font
        ws['A7'].value = "Development Action (70%)"
        ws['A7'].font = header_font
        ws['B7'] = res['Tip 1 (70%)']
        ws['B7'].font = body_font
        ws['B7'].alignment = left_align
        ws.row_dimensions[7].height = 60
        ws['A8'].value = "Development Action (20%)"
        ws['A8'].font = header_font
        ws['B8'] = res['Tip 1 (20%)']
        ws['B8'].font = body_font
        ws['B8'].alignment = left_align
        ws.row_dimensions[8].height = 60
        ws.merge_cells('A10:B10')
        ws['A10'].value = f"Focus Area 2: {res['Lowest Competency 2']}"
        ws['A10'].font = header_font
        ws['A11'].value = "Development Action (70%)"
        ws['A11'].font = header_font
        ws['B11'] = res['Tip 2 (70%)']
        ws['B11'].font = body_font
        ws['B11'].alignment = left_align
        ws.row_dimensions[11].height = 60
        ws['A12'].value = "Development Action (20%)"
        ws['A12'].font = header_font
        ws['B12'] = res['Tip 2 (20%)']
        ws['B12'].font = body_font
        ws['B12'].alignment = left_align
        ws.row_dimensions[12].height = 60

    wb.save(output)
    output.seek(0)
    return output

# --- STREAMLIT UI ---

st.set_page_config(page_title="Development Plan Generator", layout="wide")
st.title("Development Plan Generator ⚙️")

st.info("Upload your Excel files below. The app will generate a formatted Excel report.")

col1, col2 = st.columns(2)

with col1:
    st.header("Candidate Data")
    uploaded_candidates_file = st.file_uploader("1. Upload Candidate Data (Excel)", type=["xlsx"])

    @st.cache_data
    def create_sample_candidate_file():
        sample_data = {
            'Candidate Name': ['John Doe', 'Jane Smith'], 'Level': ['Apply', 'Guide'],
            'Manages Stakeholders': [2.5, 4.1], 'Steers Change': [3.1, 4.2], 'Leads People': [1.8, 4.8],
            'Drives Results': [4.5, 3.2], 'Solves Challenges': [2.1, 4.3], 'Thinks Strategically': [3.3, 4.5]
        }
        df = pd.DataFrame(sample_data)
        output = BytesIO()
        df.to_excel(output, index=False, sheet_name='Candidates')
        output.seek(0)
        return output

    st.download_button(
        label="📥 Download Sample Candidate File", data=create_sample_candidate_file(),
        file_name="sample_candidate_data.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with col2:
    st.header("Tip Repository")
    uploaded_repo_files = st.file_uploader(
        "2. Upload ALL THREE Tip Repositories (Apply, Guide, Shape Excel files)", 
        type=["xlsx"], 
        accept_multiple_files=True
    )
    
    @st.cache_data
    def create_sample_repo_file(level):
        """Creates a sample repository Excel file for a given level."""
        sample_data = {
            'Competency Name': ['Manages Stakeholders', 'Manages Stakeholders', 'Leads People', 'Leads People', ''],
            '70% Development Tips': [f'70% Tip A for {level}', f'70% Tip B for {level}', f'70% Tip C for {level}', f'70% Tip D for {level}', ''],
            '20% Development Tips': [f'20% Tip X for {level}', f'20% Tip Y for {level}', f'20% Tip Z for {level}', f'20% Tip W for {level}', '']
        }
        df = pd.DataFrame(sample_data)
        output = BytesIO()
        df.to_excel(output, index=False, sheet_name=level)
        output.seek(0)
        return output

    st.download_button(
        label="📥 Download Sample 'Apply' Repo", data=create_sample_repo_file('Apply'),
        file_name="sample_repository_apply.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='apply'
    )
    st.download_button(
        label="📥 Download Sample 'Guide' Repo", data=create_sample_repo_file('Guide'),
        file_name="sample_repository_guide.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='guide'
    )
    st.download_button(
        label="📥 Download Sample 'Shape' Repo", data=create_sample_repo_file('Shape'),
        file_name="sample_repository_shape.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key='shape'
    )

st.markdown("---")

if st.button("Generate Development Reports", type="primary"):
    if uploaded_candidates_file is None:
        st.error("⚠️ Please upload the Candidate Data Excel file.")
    elif len(uploaded_repo_files) != 3:
        st.error(f"⚠️ Please upload exactly 3 repository Excel files. You have uploaded {len(uploaded_repo_files)}.")
    else:
        with st.spinner('Processing...'):
            try:
                candidates_df = pd.read_excel(uploaded_candidates_file)
                
                repos = {}
                for f in uploaded_repo_files:
                    if 'apply' in f.name.lower():
                        repos['Apply'] = pd.read_excel(f)
                    elif 'guide' in f.name.lower():
                        repos['Guide'] = pd.read_excel(f)
                    elif 'shape' in f.name.lower():
                        repos['Shape'] = pd.read_excel(f)
                
                if len(repos) != 3:
                    st.error("❌ Could not identify 'Apply', 'Guide', and 'Shape' files from the filenames. Please ensure your uploaded Excel files contain these words.")
                    st.stop()
                
                st.write("✅ Repositories loaded successfully!")

                COMPETENCIES = [
                    'Manages Stakeholders', 'Steers Change', 'Leads People', 
                    'Drives Results', 'Solves Challenges', 'Thinks Strategically'
                ]
                final_results = []
                
                for index, row in candidates_df.iterrows():
                    level = row['Level']
                    if level not in repos:
                        st.warning(f"Skipping candidate {row['Candidate Name']}: level '{level}' not found.")
                        continue
                    
                    repo_df = repos[level].copy() 
                    low_comp_1, low_comp_2 = find_lowest_competencies(row, COMPETENCIES)
                    
                    if not low_comp_1 or not low_comp_2:
                        st.warning(f"Skipping candidate {row['Candidate Name']}: could not determine two lowest competencies.")
                        continue
                    
                    tip1_70, tip1_20 = get_random_tips(repo_df, low_comp_1)
                    tip2_70, tip2_20 = get_random_tips(repo_df, low_comp_2)
                    
                    final_results.append({
                        "Candidate Name": row['Candidate Name'], "Level": level,
                        "Lowest Competency 1": low_comp_1, "Tip 1 (70%)": tip1_70, "Tip 1 (20%)": tip1_20,
                        "Lowest Competency 2": low_comp_2, "Tip 2 (70%)": tip2_70, "Tip 2 (20%)": tip2_20,
                    })

                if final_results:
                    excel_data = generate_formatted_excel(final_results)
                    st.success("✅ Success! Your development plans are ready for download.")
                    st.download_button(
                        label="📥 Download Formatted Excel Report", data=excel_data,
                        file_name="Development_Plans.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("No candidates were processed. Please check your input files.")

            except Exception as e:
                st.error(f"An unexpected error occurred. Please check that your Excel files are formatted correctly. Error details: {e}")
